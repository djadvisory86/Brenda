"""
🏪 Rental Payment Manager Bot — Enhanced Edition
Всі фічі: GPS, прогноз, Google Sheets, авто-нагадування,
швидка оплата, звіти на пошту, шаблони, підвищення оренди,
договори, категорії, порівняння місяців
"""

import logging, os, json, asyncio
from datetime import datetime, date, time as dtime, timedelta
from calendar import monthrange
from io import BytesIO

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    ReplyKeyboardMarkup, KeyboardButton, InputMediaPhoto
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, ContextTypes, filters
)
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Config ──────────────────────────────────────────────────
BOT_TOKEN   = os.environ.get("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")
SHEETS_URL  = os.environ.get("SHEETS_WEBHOOK", "")   # Google Apps Script webhook
EMAIL_CHAT  = os.environ.get("REPORT_CHAT_ID", "")   # Telegram chat ID for monthly report
DB_FILE     = "rental_data.json"

MONTHS_UA = ["Січень","Лютий","Березень","Квітень","Травень","Червень",
             "Липень","Серпень","Вересень","Жовтень","Листопад","Грудень"]

CATEGORIES = ["🏬 ТЦ", "🏪 Ринок", "🏢 Офіс", "🛣️ Вулиця", "🏭 Склад", "📦 Інше"]

# ConvHandler states
(
    POINT_NAME, POINT_CATEGORY, POINT_ADDRESS, POINT_RENT,
    POINT_CONTACT, POINT_PHONE, POINT_DEADLINE, POINT_REMINDER_DAYS,
    PAY_AMOUNT, PAY_NOTE, PAY_DATE, PAY_RECEIPT,
    RENT_INCREASE_AMOUNT, RENT_INCREASE_NOTE,
    DOC_NAME, DOC_FILE,
    QUICK_PAY_CONFIRM,
    TEMPLATE_NAME,
) = range(18)

# ── Database ────────────────────────────────────────────────
def load_db() -> dict:
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"users": {}, "points": {}, "payments": {}, "documents": {}, "templates": {}}

def save_db(db: dict):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)

def get_user_points(db, user_id) -> list:
    uid = str(user_id)
    return [{**p, "id": k} for k, p in db["points"].items()
            if p.get("owner") == uid or uid in p.get("partners", [])]

def payment_key(pid, y, m): return f"{pid}_{y}_{m}"
def get_payment(db, pid, y, m): return db["payments"].get(payment_key(pid, y, m))
def now_ym(): n = datetime.now(); return n.year, n.month
def fmt(v): return f"{v:,.0f} ₴".replace(",", " ")
def days_to_eom():
    n = date.today(); return monthrange(n.year, n.month)[1] - n.day

def get_user(db, uid):
    k = str(uid)
    if k not in db["users"]:
        db["users"][k] = {"partners": [], "role": "owner", "reminder_days": 3}
    return db["users"][k]

# ── Google Sheets sync ──────────────────────────────────────
async def sync_to_sheets(payment_data: dict):
    """Надсилає дані оплати до Google Sheets через Apps Script webhook"""
    if not SHEETS_URL:
        return
    try:
        import urllib.request
        data = json.dumps(payment_data).encode()
        req = urllib.request.Request(SHEETS_URL, data=data,
              headers={"Content-Type": "application/json"}, method="POST")
        urllib.request.urlopen(req, timeout=5)
        logger.info("Synced to Google Sheets")
    except Exception as e:
        logger.warning(f"Sheets sync failed: {e}")

# ── Keyboards ───────────────────────────────────────────────
def main_kb():
    return ReplyKeyboardMarkup([
        ["🏪 Мої точки", "⚡ Оплатити все"],
        ["📊 Статистика", "📅 Місяць"],
        ["📈 Прогноз", "🔔 Нагадування"],
        ["⚙️ Налаштування", "❓ Допомога"],
    ], resize_keyboard=True)

def months_kb(year, month):
    btns = []
    row = []
    for i in range(12):
        m = i + 1
        mark = "✅" if m == month else ""
        row.append(InlineKeyboardButton(f"{mark}{MONTHS_UA[i][:3]}", callback_data=f"selmonth_{year}_{m}"))
        if len(row) == 3: btns.append(row); row = []
    if row: btns.append(row)
    btns.append([
        InlineKeyboardButton("◀️", callback_data=f"selyear_{year-1}_{month}"),
        InlineKeyboardButton(f"📅 {year}", callback_data="noop"),
        InlineKeyboardButton("▶️", callback_data=f"selyear_{year+1}_{month}"),
    ])
    return InlineKeyboardMarkup(btns)

def category_kb():
    btns = [[InlineKeyboardButton(c, callback_data=f"cat_{c}")] for c in CATEGORIES]
    btns.append([InlineKeyboardButton("⏭️ Пропустити", callback_data="cat_skip")])
    return InlineKeyboardMarkup(btns)

# ── Helpers ─────────────────────────────────────────────────
def check_debts(db, user_id) -> list:
    debts = []
    points = get_user_points(db, user_id)
    n = datetime.now()
    for i in range(1, 4):
        m = n.month - i; y = n.year
        if m <= 0: m += 12; y -= 1
        for p in points:
            if not get_payment(db, p["id"], y, m)?.get("paid"):
                debts.append({"point": p, "year": y, "month": m})
    return debts

def annual_forecast(db, user_id) -> dict:
    """Прогноз витрат на рік з урахуванням підвищень оренди"""
    points = get_user_points(db, user_id)
    year, month = now_ym()

    # Скільки вже заплатили цього року
    paid_ytd = 0
    for m in range(1, month + 1):
        for p in points:
            pay = get_payment(db, p["id"], year, m)
            if pay and pay.get("paid"):
                paid_ytd += pay["amount"]

    # Прогноз на решту місяців
    monthly_total = sum(p["rent"] for p in points)
    remaining_months = 12 - month
    forecast_rest = monthly_total * remaining_months

    # Річний план (якщо б платили завжди)
    annual_plan = monthly_total * 12

    # Середня оплата за місяць
    months_counted = max(month, 1)
    avg_monthly = paid_ytd / months_counted

    return {
        "paid_ytd": paid_ytd,
        "forecast_rest": forecast_rest,
        "annual_plan": annual_plan,
        "annual_forecast": paid_ytd + forecast_rest,
        "monthly_total": monthly_total,
        "avg_monthly": avg_monthly,
        "remaining_months": remaining_months,
        "savings_vs_plan": annual_plan - (paid_ytd + forecast_rest),
        "points_count": len(points),
    }

def compare_months(db, user_id, y1, m1, y2, m2) -> dict:
    """Порівняння двох місяців"""
    points = get_user_points(db, user_id)
    def month_stats(y, m):
        total = sum(p["rent"] for p in points)
        paid_pts = [p for p in points if get_payment(db, p["id"], y, m)?.get("paid")]
        paid_sum = sum(get_payment(db, p["id"], y, m)["amount"] for p in paid_pts)
        return {"total": total, "paid": paid_sum, "count": len(paid_pts),
                "pct": int(paid_sum / total * 100) if total else 0}
    s1 = month_stats(y1, m1)
    s2 = month_stats(y2, m2)
    return {"prev": s1, "curr": s2,
            "diff_sum": s2["paid"] - s1["paid"],
            "diff_pct": s2["pct"] - s1["pct"]}

# ── /start ──────────────────────────────────────────────────
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    u = update.effective_user
    get_user(db, u.id); save_db(db)
    debts = check_debts(db, u.id)
    debt_warn = f"\n\n🚨 Знайдено <b>{len(debts)}</b> неоплачених місяців!" if debts else ""
    await update.message.reply_html(
        f"👋 Привіт, <b>{u.first_name}</b>! Твій помічник оренди готовий.\n\n"
        f"<b>Можливості:</b>\n"
        f"🏪 Точки з GPS, категоріями, договорами\n"
        f"💰 Оплата з фото чека\n"
        f"📈 Прогноз витрат на рік\n"
        f"🔔 Авто-нагадування до дедлайну\n"
        f"📊 Порівняння місяців, аналітика\n"
        f"🔄 Синхронізація з Google Sheets\n"
        f"⚡ Швидка оплата всіх точок одним кліком"
        f"{debt_warn}",
        reply_markup=main_kb()
    )

# ── Мої точки ───────────────────────────────────────────────
async def cmd_points(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    uid = update.effective_user.id
    points = get_user_points(db, uid)
    year, month = now_ym()

    if not points:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("➕ Додати точку", callback_data="add_point")]])
        await update.message.reply_text("Точок ще немає. Додай першу! 👆", reply_markup=kb)
        return

    paid   = [p for p in points if get_payment(db, p["id"], year, month)?.get("paid")]
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]
    total  = sum(p["rent"] for p in points)
    paid_sum = sum(get_payment(db, p["id"], year, month)["amount"]
                   for p in paid if get_payment(db, p["id"], year, month))

    # Групуємо по категоріях
    cats = {}
    for p in points:
        c = p.get("category", "📦 Інше")
        cats.setdefault(c, []).append(p)

    text = f"🏪 <b>{MONTHS_UA[month-1]} {year}</b>\n\n"
    for cat, pts in cats.items():
        text += f"<b>{cat}</b>\n"
        for p in pts:
            pay = get_payment(db, p["id"], year, month)
            if pay and pay.get("paid"):
                text += f"  ✅ {p['name']} — {fmt(pay['amount'])} ({pay['date']})\n"
            else:
                dl = f" ⏰до {p.get('deadline','')}го" if p.get("deadline") else ""
                text += f"  ❌ {p['name']} — {fmt(p['rent'])}{dl}\n"
        text += "\n"

    pct = int(paid_sum / total * 100) if total else 0
    bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
    text += f"<code>{bar}</code> {pct}%\n"
    text += f"💵 {fmt(paid_sum)} / {fmt(total)}  •  Залишок: {fmt(total-paid_sum)}"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 Оплатити", callback_data=f"choose_pay_{year}_{month}"),
         InlineKeyboardButton("⚡ Всі одразу", callback_data=f"quickpay_{year}_{month}")],
        [InlineKeyboardButton("📋 Деталі", callback_data=f"points_detail_{year}_{month}"),
         InlineKeyboardButton("➕ Додати", callback_data="add_point")],
    ])
    await update.message.reply_html(text, reply_markup=kb)

# ── ⚡ Швидка оплата всіх ────────────────────────────────────
async def cb_quickpay(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[1]), int(parts[2])
    db = load_db()
    uid = query.from_user.id
    points = get_user_points(db, uid)
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]

    if not unpaid:
        await query.edit_message_text("✅ Всі точки вже оплачені!")
        return

    total = sum(p["rent"] for p in unpaid)
    text = f"⚡ <b>Швидка оплата</b>\n{MONTHS_UA[month-1]} {year}\n\n"
    for p in unpaid:
        text += f"  • {p['name']} — {fmt(p['rent'])}\n"
    text += f"\n💵 Загалом: <b>{fmt(total)}</b>"

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Підтвердити всі", callback_data=f"quickpay_confirm_{year}_{month}"),
        InlineKeyboardButton("❌ Скасувати", callback_data="cancel"),
    ]])
    await query.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

async def cb_quickpay_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer("⏳ Фіксую...")
    parts = query.data.split("_")
    year, month = int(parts[2]), int(parts[3])
    db = load_db()
    uid = query.from_user.id
    points = get_user_points(db, uid)
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]

    today_str = datetime.now().strftime("%d.%m.%Y")
    count = 0
    for p in unpaid:
        key = payment_key(p["id"], year, month)
        pay_data = {"paid": True, "amount": p["rent"], "note": "Швидка оплата",
                    "date": today_str, "ts": int(datetime.now().timestamp()), "user": str(uid)}
        db["payments"][key] = pay_data
        count += 1
        await sync_to_sheets({"point": p["name"], "month": month, "year": year, **pay_data})

    save_db(db)
    total = sum(p["rent"] for p in unpaid)
    await query.edit_message_text(
        f"⚡ <b>Готово!</b> Оплачено {count} точок\n💵 Загалом: {fmt(total)}\n📅 {today_str}",
        parse_mode="HTML"
    )

# ── Оплата з фото чека ──────────────────────────────────────
async def cb_pay_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    pid, year, month = parts[1], int(parts[2]), int(parts[3])
    db = load_db()
    point = db["points"].get(pid)
    if not point:
        await query.edit_message_text("❌ Точку не знайдено.")
        return ConversationHandler.END

    ctx.user_data["pay"] = {"pid": pid, "year": year, "month": month, "default_amount": point["rent"]}
    await query.message.reply_html(
        f"💰 <b>{point['name']}</b> — {MONTHS_UA[month-1]} {year}\n\n"
        f"Введи суму або /skip для стандартної {fmt(point['rent'])}:"
    )
    return PAY_AMOUNT

async def pay_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text == "/skip":
        ctx.user_data["pay"]["amount"] = ctx.user_data["pay"]["default_amount"]
    else:
        try: ctx.user_data["pay"]["amount"] = float(update.message.text.replace(" ","").replace(",","."))
        except ValueError:
            await update.message.reply_text("⚠️ Тільки число або /skip"); return PAY_AMOUNT
    await update.message.reply_text("📝 Нотатка (квитанція №, коментар) або /skip:")
    return PAY_NOTE

async def pay_note(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        ctx.user_data["pay"]["note"] = update.message.text.strip()
    await update.message.reply_text(
        f"📅 Дата ({datetime.now().strftime('%d.%m.%Y')}) або /skip:"
    )
    return PAY_DATE

async def pay_date(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    pd = ctx.user_data["pay"]
    pd["date"] = datetime.now().strftime("%d.%m.%Y") if update.message.text == "/skip" else update.message.text.strip()
    await update.message.reply_text(
        "📸 Надішли фото чека/квитанції або /skip:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⏭️ Пропустити", callback_data="skip_receipt")]])
    )
    return PAY_RECEIPT

async def pay_receipt_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    pd = ctx.user_data["pay"]
    if update.message.photo:
        pd["receipt_file_id"] = update.message.photo[-1].file_id
    return await _save_payment(update, ctx)

async def pay_receipt_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await _save_payment(update, ctx)

async def cb_skip_receipt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    return await _save_payment(update, ctx, from_callback=True)

async def _save_payment(update, ctx, from_callback=False):
    pd = ctx.user_data["pay"]
    db = load_db()
    point = db["points"][pd["pid"]]
    key = payment_key(pd["pid"], pd["year"], pd["month"])
    pay_record = {
        "paid": True,
        "amount": pd["amount"],
        "note": pd.get("note", ""),
        "date": pd["date"],
        "ts": int(datetime.now().timestamp()),
        "user": str((update.callback_query or update).from_user.id),
        "receipt_file_id": pd.get("receipt_file_id", ""),
    }
    db["payments"][key] = pay_record
    save_db(db)

    await sync_to_sheets({"point": point["name"], "month": pd["month"],
                          "year": pd["year"], **pay_record})

    receipt_info = " 📸 Фото збережено" if pd.get("receipt_file_id") else ""
    text = (f"✅ <b>Оплачено!</b>\n🏪 {point['name']}\n"
            f"💵 {fmt(pd['amount'])}\n📅 {pd['date']}"
            + (f"\n📝 {pd.get('note','')}" if pd.get('note') else "")
            + receipt_info)

    msg = update.callback_query.message if from_callback else update.message
    await msg.reply_html(text, reply_markup=main_kb())
    ctx.user_data.clear()
    return ConversationHandler.END

# ── Деталі + GPS ────────────────────────────────────────────
async def cb_points_detail(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[2]), int(parts[3])
    db = load_db()
    points = get_user_points(db, query.from_user.id)

    text = f"📋 <b>Деталі — {MONTHS_UA[month-1]} {year}</b>\n\n"
    btns = []
    for p in points:
        pay = get_payment(db, p["id"], year, month)
        status = "✅" if pay and pay.get("paid") else "❌"
        text += f"{status} <b>{p['name']}</b>  {p.get('category','')}\n"
        text += f"   📍 {p.get('address', '—')}\n"
        text += f"   💵 Оренда: {fmt(p['rent'])}\n"
        if p.get("contact"):
            text += f"   👤 {p['contact']}"
            if p.get("phone"): text += f"  📞 {p['phone']}"
            text += "\n"
        if p.get("deadline"):
            text += f"   ⏰ Дедлайн: до {p['deadline']}-го\n"
        if p.get("lat"):
            text += f"   🗺️ GPS збережено\n"
        if pay and pay.get("paid"):
            text += f"   ✅ {pay['date']}: {fmt(pay['amount'])}"
            if pay.get("note"): text += f" · {pay['note']}"
            if pay.get("receipt_file_id"): text += " 📸"
            text += "\n"

        # Кнопки для точки
        row = []
        if pay and pay.get("paid"):
            row.append(InlineKeyboardButton(f"↩️ {p['name']}", callback_data=f"unpay_{p['id']}_{year}_{month}"))
        else:
            row.append(InlineKeyboardButton(f"💰 {p['name']}", callback_data=f"pay_{p['id']}_{year}_{month}"))
        if p.get("lat"):
            row.append(InlineKeyboardButton("🗺️", callback_data=f"show_gps_{p['id']}"))
        btns.append(row)
        text += "\n"

    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(btns), parse_mode="HTML")

async def cb_show_gps(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    pid = query.data.split("_")[2]
    db = load_db()
    p = db["points"].get(pid, {})
    if p.get("lat"):
        await query.message.reply_location(latitude=p["lat"], longitude=p["lon"])
        await query.message.reply_text(f"📍 <b>{p['name']}</b>\n{p.get('address','')}", parse_mode="HTML")

# ── Скасувати оплату ────────────────────────────────────────
async def cb_unpay(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    pid, year, month = parts[1], int(parts[2]), int(parts[3])
    db = load_db()
    key = payment_key(pid, year, month)
    db["payments"].pop(key, None)
    save_db(db)
    name = db["points"].get(pid, {}).get("name", "")
    await query.edit_message_text(f"↩️ Оплату <b>{name}</b> за {MONTHS_UA[month-1]} {year} скасовано.", parse_mode="HTML")

# ── Вибір точки для оплати ──────────────────────────────────
async def cb_choose_pay(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[2]), int(parts[3])
    db = load_db()
    points = get_user_points(db, query.from_user.id)
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]
    if not unpaid:
        await query.edit_message_text("🎉 Всі точки оплачені!")
        return
    btns = [[InlineKeyboardButton(
        f"{p.get('category','')} {p['name']} — {fmt(p['rent'])}",
        callback_data=f"pay_{p['id']}_{year}_{month}"
    )] for p in unpaid]
    btns.append([InlineKeyboardButton("❌ Скасувати", callback_data="cancel")])
    await query.edit_message_text(
        f"💰 <b>Оплата — {MONTHS_UA[month-1]} {year}</b>\nВибери точку:",
        reply_markup=InlineKeyboardMarkup(btns), parse_mode="HTML"
    )

# ── ConvHandler: Додати точку (з GPS і категорією) ───────────
async def cb_add_point(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if q: await q.answer(); msg = q.message
    else: msg = update.message
    await msg.reply_text(
        "🏪 <b>Нова точка</b>\n\nВибери категорію:",
        reply_markup=category_kb(), parse_mode="HTML"
    )
    ctx.user_data["new_point"] = {}
    return POINT_CATEGORY

async def point_category(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    if query.data != "cat_skip":
        ctx.user_data["new_point"]["category"] = query.data.replace("cat_", "")
    await query.message.reply_text("✏️ Введи назву точки:")
    return POINT_NAME

async def point_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_point"]["name"] = update.message.text.strip()
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📍 Надіслати геолокацію", callback_data="send_location_hint"),
        InlineKeyboardButton("⏭️ Пропустити", callback_data="skip_location"),
    ]])
    await update.message.reply_text(
        "🗺️ Надішли геолокацію точки або введи адресу текстом\n"
        "(або /skip щоб пропустити):",
        reply_markup=kb
    )
    return POINT_ADDRESS

async def point_address_location(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Отримуємо геолокацію"""
    loc = update.message.location
    ctx.user_data["new_point"]["lat"] = loc.latitude
    ctx.user_data["new_point"]["lon"] = loc.longitude
    await update.message.reply_text("✅ GPS збережено! Введи текстову адресу або /skip:")
    return POINT_ADDRESS

async def point_address_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        ctx.user_data["new_point"]["address"] = update.message.text.strip()
    await update.message.reply_text("💵 Сума щомісячної оренди (число):")
    return POINT_RENT

async def point_rent(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    try:
        ctx.user_data["new_point"]["rent"] = float(update.message.text.replace(" ","").replace(",","."))
    except ValueError:
        await update.message.reply_text("⚠️ Тільки число! Наприклад: 3500"); return POINT_RENT
    await update.message.reply_text("👤 Ім'я орендодавця або /skip:")
    return POINT_CONTACT

async def point_contact(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        ctx.user_data["new_point"]["contact"] = update.message.text.strip()
    await update.message.reply_text("📞 Телефон або /skip:")
    return POINT_PHONE

async def point_phone(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        ctx.user_data["new_point"]["phone"] = update.message.text.strip()
    await update.message.reply_text("📅 До якого числа платити? (наприклад: 5) або /skip:")
    return POINT_DEADLINE

async def point_deadline(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        ctx.user_data["new_point"]["deadline"] = update.message.text.strip()
    await update.message.reply_text(
        "🔔 За скільки днів нагадувати до дедлайну?\n(наприклад: 3) або /skip:"
    )
    return POINT_REMINDER_DAYS

async def point_reminder_days(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.message.text != "/skip":
        try: ctx.user_data["new_point"]["reminder_days"] = int(update.message.text)
        except ValueError: pass

    db = load_db()
    uid = str(update.effective_user.id)
    np = ctx.user_data["new_point"]
    pid = str(int(datetime.now().timestamp() * 1000))
    np.update({"owner": uid, "partners": [], "created": datetime.now().strftime("%d.%m.%Y"),
               "rent_history": [{"rent": np["rent"], "date": np["created"] if "created" in np else datetime.now().strftime("%d.%m.%Y"), "note": "Початкова оренда"}]})
    db["points"][pid] = np
    save_db(db)

    gps_info = " 🗺️ GPS збережено" if np.get("lat") else ""
    await update.message.reply_html(
        f"✅ <b>Точку додано!</b>\n\n"
        f"{np.get('category','')} <b>{np['name']}</b>{gps_info}\n"
        f"📍 {np.get('address','—')}\n"
        f"💵 {fmt(np['rent'])}/міс\n"
        f"👤 {np.get('contact','—')}\n"
        f"📞 {np.get('phone','—')}\n"
        f"⏰ Дедлайн: {np.get('deadline','—')}-го  🔔 За {np.get('reminder_days',3)} дні",
        reply_markup=main_kb()
    )
    ctx.user_data.clear()
    return ConversationHandler.END

# ── 📈 Прогноз ───────────────────────────────────────────────
async def cmd_forecast(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    uid = update.effective_user.id
    points = get_user_points(db, uid)
    year, month = now_ym()

    if not points:
        await update.message.reply_text("📈 Спочатку додай точки!")
        return

    f = annual_forecast(db, uid)

    # Графік по місяцях
    monthly_bars = ""
    for m in range(1, 13):
        m_paid = sum(
            get_payment(db, p["id"], year, m)["amount"]
            for p in points
            if get_payment(db, p["id"], year, m)?.get("paid")
        )
        m_total = f["monthly_total"]
        pct = int(m_paid / m_total * 100) if m_total else 0
        bar = "█" * (pct // 20) + "░" * (5 - pct // 20)
        marker = " ◀" if m == month else ""
        monthly_bars += f"{MONTHS_UA[m-1][:3]} <code>{bar}</code> {fmt(m_paid)}{marker}\n"

    # Динаміка оренди по точках
    rent_changes = ""
    for p in points:
        history = p.get("rent_history", [])
        if len(history) > 1:
            first = history[0]["rent"]
            last  = history[-1]["rent"]
            delta = last - first
            sign  = "📈" if delta > 0 else "📉"
            rent_changes += f"  {sign} {p['name']}: {fmt(first)} → {fmt(last)} ({'+' if delta > 0 else ''}{fmt(delta)})\n"

    text = (
        f"📈 <b>Прогноз витрат — {year} рік</b>\n\n"
        f"<b>По місяцях:</b>\n{monthly_bars}\n"
        f"<b>Підсумок:</b>\n"
        f"💵 Сплачено з початку року: <b>{fmt(f['paid_ytd'])}</b>\n"
        f"🔮 Прогноз до кінця року: <b>{fmt(f['forecast_rest'])}</b>\n"
        f"📊 Разом за рік: <b>{fmt(f['annual_forecast'])}</b>\n"
        f"📋 План на рік: {fmt(f['annual_plan'])}\n"
        f"📌 Середнє/міс: {fmt(f['avg_monthly'])}\n"
        f"🏪 Активних точок: {f['points_count']}\n"
    )
    if rent_changes:
        text += f"\n<b>Зміни оренди:</b>\n{rent_changes}"

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("📉 Порівняти місяці", callback_data=f"compare_{year}_{month}"),
        InlineKeyboardButton("📤 Excel звіт", callback_data=f"export_{year}_{month}"),
    ]])
    await update.message.reply_html(text, reply_markup=kb)

# ── 📉 Порівняння місяців ────────────────────────────────────
async def cb_compare(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[1]), int(parts[2])

    # Порівнюємо з попереднім місяцем
    prev_m = month - 1; prev_y = year
    if prev_m <= 0: prev_m = 12; prev_y -= 1

    db = load_db()
    uid = query.from_user.id
    cmp = compare_months(db, uid, prev_y, prev_m, year, month)
    points = get_user_points(db, uid)

    curr = cmp["curr"]; prev = cmp["prev"]
    diff = cmp["diff_sum"]
    diff_str = f"{'📈 +' if diff >= 0 else '📉 '}{fmt(abs(diff))}"

    text = (
        f"📉 <b>Порівняння місяців</b>\n\n"
        f"{'─'*30}\n"
        f"<b>{MONTHS_UA[prev_m-1]} {prev_y}</b>\n"
        f"  💵 Сплачено: {fmt(prev['paid'])} ({prev['pct']}%)\n"
        f"  ✅ Точок: {prev['count']}/{len(points)}\n\n"
        f"<b>{MONTHS_UA[month-1]} {year}</b>\n"
        f"  💵 Сплачено: {fmt(curr['paid'])} ({curr['pct']}%)\n"
        f"  ✅ Точок: {curr['count']}/{len(points)}\n"
        f"{'─'*30}\n"
        f"Різниця: {diff_str}\n"
    )

    # По точках
    text += "\n<b>По точках:</b>\n"
    for p in points:
        p1 = get_payment(db, p["id"], prev_y, prev_m)
        p2 = get_payment(db, p["id"], year, month)
        s1 = "✅" if p1 and p1.get("paid") else "❌"
        s2 = "✅" if p2 and p2.get("paid") else "❌"
        text += f"  {p['name']}: {s1}→{s2}\n"

    await query.edit_message_text(text, parse_mode="HTML")

# ── 📈 Підвищення оренди ─────────────────────────────────────
async def cb_rent_increase_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    db = load_db()
    points = get_user_points(db, query.from_user.id)
    btns = [[InlineKeyboardButton(
        f"📈 {p['name']} (зараз {fmt(p['rent'])})",
        callback_data=f"increase_{p['id']}"
    )] for p in points]
    btns.append([InlineKeyboardButton("← Назад", callback_data="settings_back")])
    await query.edit_message_text("📈 Вибери точку для зміни оренди:", reply_markup=InlineKeyboardMarkup(btns))

async def cb_rent_increase_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    pid = query.data.split("_")[1]
    db = load_db()
    p = db["points"].get(pid, {})
    ctx.user_data["increase"] = {"pid": pid}
    await query.message.reply_html(
        f"📈 <b>Зміна оренди: {p.get('name','')}</b>\n"
        f"Поточна: {fmt(p.get('rent',0))}\n\n"
        f"Введи нову суму:"
    )
    return RENT_INCREASE_AMOUNT

async def rent_increase_amount(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    try:
        ctx.user_data["increase"]["new_rent"] = float(update.message.text.replace(" ","").replace(",","."))
    except ValueError:
        await update.message.reply_text("⚠️ Тільки число!"); return RENT_INCREASE_AMOUNT
    await update.message.reply_text("📝 Причина підвищення або /skip:")
    return RENT_INCREASE_NOTE

async def rent_increase_note(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    inc = ctx.user_data["increase"]
    note = "" if update.message.text == "/skip" else update.message.text.strip()
    db = load_db()
    p = db["points"][inc["pid"]]
    old_rent = p["rent"]
    p["rent"] = inc["new_rent"]
    if "rent_history" not in p: p["rent_history"] = [{"rent": old_rent, "date": p.get("created",""), "note": ""}]
    p["rent_history"].append({"rent": inc["new_rent"], "date": datetime.now().strftime("%d.%m.%Y"), "note": note})
    save_db(db)
    delta = inc["new_rent"] - old_rent
    await update.message.reply_html(
        f"✅ <b>Оренду оновлено!</b>\n"
        f"🏪 {p['name']}\n"
        f"💵 {fmt(old_rent)} → {fmt(inc['new_rent'])}\n"
        f"{'📈 +' if delta > 0 else '📉 '}{fmt(abs(delta))} на місяць\n"
        f"📅 {datetime.now().strftime('%d.%m.%Y')}" +
        (f"\n📝 {note}" if note else ""),
        reply_markup=main_kb()
    )
    ctx.user_data.clear()
    return ConversationHandler.END

# ── 📝 Договори ──────────────────────────────────────────────
async def cb_docs_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    db = load_db()
    uid = query.from_user.id
    points = get_user_points(db, uid)
    docs = db.get("documents", {})

    text = "📝 <b>Договори та документи</b>\n\n"
    btns = []
    for p in points:
        p_docs = [d for d in docs.values() if d.get("point_id") == p["id"]]
        text += f"🏪 <b>{p['name']}</b>: {len(p_docs)} документ(ів)\n"
        for d in p_docs:
            text += f"  📄 {d['name']} ({d['date']})\n"
        btns.append([InlineKeyboardButton(f"➕ Додати до {p['name']}", callback_data=f"add_doc_{p['id']}")])

    btns.append([InlineKeyboardButton("← Назад", callback_data="settings_back")])
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(btns), parse_mode="HTML")

async def cb_add_doc_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    pid = query.data.split("_")[2]
    ctx.user_data["new_doc"] = {"point_id": pid}
    db = load_db()
    pname = db["points"].get(pid, {}).get("name", "")
    await query.message.reply_text(f"📝 Документ для '{pname}'\n\nВведи назву документа:")
    return DOC_NAME

async def doc_name(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data["new_doc"]["name"] = update.message.text.strip()
    await update.message.reply_text("📎 Надішли файл (PDF, фото) або /skip:")
    return DOC_FILE

async def doc_file(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    nd = ctx.user_data["new_doc"]
    if update.message.document:
        nd["file_id"] = update.message.document.file_id
        nd["file_type"] = "document"
    elif update.message.photo:
        nd["file_id"] = update.message.photo[-1].file_id
        nd["file_type"] = "photo"
    nd["date"] = datetime.now().strftime("%d.%m.%Y")
    db = load_db()
    if "documents" not in db: db["documents"] = {}
    doc_id = str(int(datetime.now().timestamp() * 1000))
    db["documents"][doc_id] = nd
    save_db(db)
    await update.message.reply_html(
        f"✅ <b>Документ збережено!</b>\n"
        f"📄 {nd['name']}\n📅 {nd['date']}",
        reply_markup=main_kb()
    )
    ctx.user_data.clear()
    return ConversationHandler.END

async def doc_file_skip(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await doc_file(update, ctx)

# ── 🔔 Нагадування ───────────────────────────────────────────
async def cmd_remind(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    uid = update.effective_user.id
    points = get_user_points(db, uid)
    year, month = now_ym()
    days = days_to_eom()
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]
    debts = check_debts(db, uid)

    text = "🔔 <b>Нагадування</b>\n\n"
    if days <= 5:
        text += f"⚠️ До кінця місяця <b>{days} дн.</b>!\n\n"
    if unpaid:
        text += f"❌ <b>Не оплачено ({len(unpaid)}):</b>\n"
        for p in unpaid:
            text += f"  • {p['name']} — {fmt(p['rent'])}"
            if p.get("deadline"): text += f" (до {p['deadline']}-го)"
            text += "\n"
    else:
        text += "✅ Все оплачено цього місяця!\n"
    if debts:
        total_debt = sum(d["point"]["rent"] for d in debts)
        text += f"\n🚨 <b>Борги ({len(debts)} — ~{fmt(total_debt)}):</b>\n"
        for d in debts:
            text += f"  • {d['point']['name']} — {MONTHS_UA[d['month']-1]} {d['year']}\n"

    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("💰 Оплатити", callback_data=f"choose_pay_{year}_{month}"),
        InlineKeyboardButton("⚡ Всі одразу", callback_data=f"quickpay_{year}_{month}"),
    ]] if unpaid else [])
    await update.message.reply_html(text, reply_markup=kb)

# ── Авто-нагадування (job scheduler) ────────────────────────
async def job_daily_reminder(ctx: ContextTypes.DEFAULT_TYPE):
    """Щоденна перевірка: нагадати якщо до дедлайну <= reminder_days"""
    db = load_db()
    year, month = now_ym()
    today_day = date.today().day

    for uid_str, user_data in db["users"].items():
        points = get_user_points(db, int(uid_str))
        for p in points:
            deadline = p.get("deadline")
            reminder_days = p.get("reminder_days", user_data.get("reminder_days", 3))
            if not deadline: continue
            try:
                deadline_day = int(deadline)
                days_left = deadline_day - today_day
                if days_left == reminder_days:
                    pay = get_payment(db, p["id"], year, month)
                    if not pay or not pay.get("paid"):
                        await ctx.bot.send_message(
                            chat_id=int(uid_str),
                            text=(
                                f"🔔 <b>Нагадування!</b>\n\n"
                                f"🏪 {p['name']}\n"
                                f"💵 {fmt(p['rent'])}\n"
                                f"⏰ Дедлайн через <b>{days_left} дні</b> ({deadline_day}-го числа)"
                            ),
                            parse_mode="HTML",
                            reply_markup=InlineKeyboardMarkup([[
                                InlineKeyboardButton("💰 Оплатити зараз", callback_data=f"pay_{p['id']}_{year}_{month}")
                            ]])
                        )
            except Exception as e:
                logger.warning(f"Reminder error for {uid_str}: {e}")

async def job_monthly_report(ctx: ContextTypes.DEFAULT_TYPE):
    """1-го числа місяця — надіслати зведений звіт"""
    if date.today().day != 1: return
    db = load_db()
    year, month = now_ym()
    prev_m = month - 1; prev_y = year
    if prev_m <= 0: prev_m = 12; prev_y -= 1

    for uid_str in db["users"]:
        points = get_user_points(db, int(uid_str))
        if not points: continue
        paid   = [p for p in points if get_payment(db, p["id"], prev_y, prev_m)?.get("paid")]
        unpaid = [p for p in points if not get_payment(db, p["id"], prev_y, prev_m)?.get("paid")]
        total  = sum(p["rent"] for p in points)
        paid_s = sum(get_payment(db, p["id"], prev_y, prev_m)["amount"] for p in paid if get_payment(db, p["id"], prev_y, prev_m))
        try:
            await ctx.bot.send_message(
                chat_id=int(uid_str),
                text=(
                    f"📊 <b>Звіт за {MONTHS_UA[prev_m-1]} {prev_y}</b>\n\n"
                    f"✅ Оплачено: {len(paid)}/{len(points)} точок\n"
                    f"💵 Сума: {fmt(paid_s)} / {fmt(total)}\n"
                    + (f"❌ Не оплачено: {', '.join(p['name'] for p in unpaid)}\n" if unpaid else "")
                    + f"\n🆕 Новий місяць — {MONTHS_UA[month-1]}! Вперед 💪"
                ),
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("📤 Повний Excel", callback_data=f"export_{prev_y}_{prev_m}")
                ]])
            )
        except Exception as e:
            logger.warning(f"Monthly report error for {uid_str}: {e}")

# ── 📊 Статистика ────────────────────────────────────────────
async def cmd_stats(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    uid = update.effective_user.id
    points = get_user_points(db, uid)
    year, month = now_ym()
    if not points:
        await update.message.reply_text("📊 Спочатку додай точки!"); return

    text = f"📊 <b>Статистика — {year}</b>\n\n"
    annual_paid = 0; annual_total = 0
    for m in range(1, month + 1):
        mt = sum(p["rent"] for p in points)
        mp = sum(get_payment(db, p["id"], year, m)["amount"]
                 for p in points if get_payment(db, p["id"], year, m)?.get("paid"))
        annual_total += mt; annual_paid += mp
        pct = int(mp / mt * 100) if mt else 0
        bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
        text += f"{MONTHS_UA[m-1][:3]} <code>{bar}</code> {pct}%\n"

    text += (f"\n💵 Сплачено: {fmt(annual_paid)}\n"
             f"📋 Заплановано: {fmt(annual_total)}\n"
             f"🎯 Ефективність: {int(annual_paid/annual_total*100) if annual_total else 0}%\n\n")

    # По категоріях
    cats = {}
    for p in points:
        c = p.get("category", "Інше")
        pay = get_payment(db, p["id"], year, month)
        cats.setdefault(c, {"paid": 0, "total": 0})
        cats[c]["total"] += p["rent"]
        if pay and pay.get("paid"): cats[c]["paid"] += pay["amount"]
    text += "<b>По категоріях (цей місяць):</b>\n"
    for c, v in cats.items():
        text += f"  {c}: {fmt(v['paid'])} / {fmt(v['total'])}\n"

    debts = check_debts(db, uid)
    if debts:
        text += f"\n🚨 Боргів: {len(debts)}\n"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📤 Excel", callback_data=f"export_{year}_{month}"),
         InlineKeyboardButton("📉 Порівняти", callback_data=f"compare_{year}_{month}")],
        [InlineKeyboardButton("📈 Прогноз", callback_data="forecast_inline")],
    ])
    await update.message.reply_html(text, reply_markup=kb)

# ── 📅 Місяць ────────────────────────────────────────────────
async def cmd_month(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    y, m = now_ym()
    await update.message.reply_text("📅 Вибери місяць:", reply_markup=months_kb(y, m))

async def cb_sel_month(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[1]), int(parts[2])
    db = load_db()
    uid = query.from_user.id
    points = get_user_points(db, uid)
    paid   = [p for p in points if get_payment(db, p["id"], year, month)?.get("paid")]
    unpaid = [p for p in points if not get_payment(db, p["id"], year, month)?.get("paid")]
    total  = sum(p["rent"] for p in points)
    paid_s = sum(get_payment(db, p["id"], year, month)["amount"] for p in paid if get_payment(db, p["id"], year, month))
    pct = int(paid_s / total * 100) if total else 0
    bar = "█" * (pct // 10) + "░" * (10 - pct // 10)
    text = (f"📅 <b>{MONTHS_UA[month-1]} {year}</b>\n\n"
            f"<code>{bar}</code> {pct}%\n"
            f"✅ {len(paid)}/{len(points)} · 💵 {fmt(paid_s)} / {fmt(total)}\n")
    if unpaid:
        text += "\n❌ " + ", ".join(p["name"] for p in unpaid)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("💰 Оплатити", callback_data=f"choose_pay_{year}_{month}"),
         InlineKeyboardButton("⚡ Всі", callback_data=f"quickpay_{year}_{month}")],
        [InlineKeyboardButton("📋 Деталі", callback_data=f"points_detail_{year}_{month}"),
         InlineKeyboardButton("📤 Excel", callback_data=f"export_{year}_{month}")],
        [InlineKeyboardButton("← Назад", callback_data="sel_month_nav")],
    ])
    await query.edit_message_text(text, reply_markup=kb, parse_mode="HTML")

async def cb_sel_year(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    parts = query.data.split("_")
    year, month = int(parts[1]), int(parts[2])
    await query.edit_message_reply_markup(reply_markup=months_kb(year, month))

async def cb_sel_month_nav(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    y, m = now_ym()
    await query.edit_message_text("📅 Вибери місяць:", reply_markup=months_kb(y, m))

# ── 📤 Excel ─────────────────────────────────────────────────
async def cb_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer("⏳ Генерую Excel...")
    parts = query.data.split("_")
    year, month = int(parts[1]), int(parts[2])
    db = load_db()
    points = get_user_points(db, query.from_user.id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTHS_UA[month-1]} {year}"

    hfill = PatternFill("solid", fgColor="1a1a2e")
    pfill = PatternFill("solid", fgColor="d4edda")
    ufill = PatternFill("solid", fgColor="f8d7da")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    bord  = Border(*[Side(style="thin")] * 0,
                   left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Звіт орендних платежів — {MONTHS_UA[month-1]} {year}"
    ws["A1"].font = Font(bold=True, size=15, color="1a1a2e")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Сформовано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, color="888888", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Категорія","Назва точки","Адреса","Орендодавець","Оренда","Статус","Дата оплати","Нотатка"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center"); c.border = bord
    ws.row_dimensions[4].height = 18

    total_rent = 0; total_paid = 0
    for row, p in enumerate(points, 5):
        pay = get_payment(db, p["id"], year, month)
        is_paid = pay and pay.get("paid")
        fill = pfill if is_paid else ufill
        total_rent += p["rent"]
        if is_paid: total_paid += pay["amount"]
        data = [p.get("category",""), p["name"], p.get("address",""),
                p.get("contact",""), p["rent"],
                "✅ Оплачено" if is_paid else "❌ Не оплачено",
                pay["date"] if is_paid else "", pay.get("note","") if is_paid else ""]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.border = bord
            c.alignment = Alignment(horizontal="left", vertical="center")
            if col == 5: c.number_format = '#,##0 [$₴]'

    lr = len(points) + 6
    for label, val, color in [("РАЗОМ:", total_rent, "000000"),
                               ("Сплачено:", total_paid, "1a7c3e"),
                               ("Залишок:", total_rent-total_paid, "c0392b")]:
        ws.cell(row=lr, column=4, value=label).font = Font(bold=True, color=color)
        c = ws.cell(row=lr, column=5, value=val)
        c.font = Font(bold=True, color=color); c.number_format = '#,##0 [$₴]'
        lr += 1

    for i, w in enumerate([12,22,25,22,14,16,14,25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"оренда_{MONTHS_UA[month-1]}_{year}.xlsx"
    await query.message.reply_document(document=buf, filename=fname,
        caption=f"📊 {MONTHS_UA[month-1]} {year} · {fmt(total_paid)} / {fmt(total_rent)}")

# ── ⚙️ Налаштування ──────────────────────────────────────────
async def cmd_settings(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    db = load_db()
    points = get_user_points(db, update.effective_user.id)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Додати точку", callback_data="add_point")],
        [InlineKeyboardButton("🗑️ Видалити точку", callback_data="delete_point_list")],
        [InlineKeyboardButton("📈 Підвищення оренди", callback_data="rent_increase_list")],
        [InlineKeyboardButton("📝 Договори", callback_data="docs_list")],
        [InlineKeyboardButton("👥 Партнери", callback_data="partners_menu")],
    ])
    await update.message.reply_html(
        f"⚙️ <b>Налаштування</b>\n\nТочок: {len(points)}\n"
        f"ID: <code>{update.effective_user.id}</code>",
        reply_markup=kb
    )

async def cb_delete_point_list(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    db = load_db()
    points = get_user_points(db, query.from_user.id)
    if not points:
        await query.edit_message_text("Точок немає."); return
    btns = [[InlineKeyboardButton(f"🗑️ {p['name']}", callback_data=f"del_confirm_{p['id']}")] for p in points]
    btns.append([InlineKeyboardButton("← Назад", callback_data="settings_back")])
    await query.edit_message_text("Вибери точку для видалення:", reply_markup=InlineKeyboardMarkup(btns))

async def cb_del_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    pid = query.data.split("_")[2]
    db = load_db()
    p = db["points"].get(pid, {})
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Видалити", callback_data=f"del_do_{pid}"),
        InlineKeyboardButton("❌ Скасувати", callback_data="settings_back"),
    ]])
    await query.edit_message_text(f"🗑️ Видалити <b>{p.get('name','')}</b>?", reply_markup=kb, parse_mode="HTML")

async def cb_del_do(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    pid = query.data.split("_")[2]
    db = load_db()
    name = db["points"].pop(pid, {}).get("name", "")
    save_db(db)
    await query.edit_message_text(f"✅ <b>{name}</b> видалено.", parse_mode="HTML")

async def cb_partners_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    uid = query.from_user.id
    await query.edit_message_html(
        f"👥 <b>Партнери</b>\n\n"
        f"Щоб додати партнера, він має:\n"
        f"1. Написати боту /start\n"
        f"2. Надіслати тобі свій ID\n"
        f"3. Ти повідомляєш йому свій ID: <code>{uid}</code>\n\n"
        f"Партнер зможе переглядати і відмічати оплати."
    )

async def cb_settings_back(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query; await query.answer()
    db = load_db()
    points = get_user_points(db, query.from_user.id)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Додати точку", callback_data="add_point")],
        [InlineKeyboardButton("🗑️ Видалити точку", callback_data="delete_point_list")],
        [InlineKeyboardButton("📈 Підвищення оренди", callback_data="rent_increase_list")],
        [InlineKeyboardButton("📝 Договори", callback_data="docs_list")],
    ])
    await query.edit_message_text(f"⚙️ <b>Налаштування</b>\nТочок: {len(points)}", reply_markup=kb, parse_mode="HTML")

# ── ❓ Допомога ──────────────────────────────────────────────
async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_html(
        "❓ <b>Як користуватись</b>\n\n"
        "🏪 <b>Мої точки</b> — статус оплат з категоріями\n"
        "⚡ <b>Оплатити все</b> — всі точки одним натиском\n"
        "📊 <b>Статистика</b> — аналітика по місяцях і категоріях\n"
        "📈 <b>Прогноз</b> — скільки заплатиш до кінця року\n"
        "🔔 <b>Нагадування</b> — борги і поточні неоплачені\n\n"
        "<b>При оплаті можна:</b>\n"
        "• Вказати іншу суму\n"
        "• Додати нотатку\n"
        "• Прикріпити фото чека 📸\n\n"
        "<b>Авто-функції:</b>\n"
        "• Нагадування за N днів до дедлайну\n"
        "• Звіт 1-го числа місяця\n"
        "• Синхронізація з Google Sheets\n\n"
        f"👤 Твій ID: <code>{update.effective_user.id}</code>"
    )

# ── Text router ──────────────────────────────────────────────
async def text_handler(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = update.message.text
    if t == "🏪 Мої точки":       await cmd_points(update, ctx)
    elif t == "⚡ Оплатити все":
        y, m = now_ym()
        fake_q = type("Q", (), {"data": f"quickpay_{y}_{m}", "answer": lambda self: None,
                                "from_user": update.effective_user, "message": update.message,
                                "edit_message_text": update.message.reply_html})()
        update.callback_query = fake_q
        await cb_quickpay(update, ctx)
    elif t == "📊 Статистика":     await cmd_stats(update, ctx)
    elif t == "📅 Місяць":         await cmd_month(update, ctx)
    elif t == "📈 Прогноз":        await cmd_forecast(update, ctx)
    elif t == "🔔 Нагадування":    await cmd_remind(update, ctx)
    elif t == "⚙️ Налаштування":  await cmd_settings(update, ctx)
    elif t == "❓ Допомога":       await cmd_help(update, ctx)

async def cb_cancel(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    if q: await q.answer(); await q.edit_message_text("❌ Скасовано.")
    ctx.user_data.clear()
    return ConversationHandler.END

async def cb_noop(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if update.callback_query: await update.callback_query.answer()

# ── MAIN ─────────────────────────────────────────────────────
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    # ConvHandler: додавання точки
    add_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cb_add_point, pattern="^add_point$"),
                      CommandHandler("add", cb_add_point)],
        states={
            POINT_CATEGORY:     [CallbackQueryHandler(point_category, pattern="^cat_")],
            POINT_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, point_name)],
            POINT_ADDRESS:      [MessageHandler(filters.LOCATION, point_address_location),
                                 MessageHandler(filters.TEXT, point_address_text),
                                 CallbackQueryHandler(lambda u,c: point_address_text(u,c), pattern="^skip_location$")],
            POINT_RENT:         [MessageHandler(filters.TEXT & ~filters.COMMAND, point_rent)],
            POINT_CONTACT:      [MessageHandler(filters.TEXT, point_contact)],
            POINT_PHONE:        [MessageHandler(filters.TEXT, point_phone)],
            POINT_DEADLINE:     [MessageHandler(filters.TEXT, point_deadline)],
            POINT_REMINDER_DAYS:[MessageHandler(filters.TEXT, point_reminder_days)],
        },
        fallbacks=[CommandHandler("cancel", cb_cancel)], per_message=False,
    )

    # ConvHandler: оплата
    pay_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cb_pay_start, pattern=r"^pay_\w+_\d+_\d+$")],
        states={
            PAY_AMOUNT:  [MessageHandler(filters.TEXT, pay_amount)],
            PAY_NOTE:    [MessageHandler(filters.TEXT, pay_note)],
            PAY_DATE:    [MessageHandler(filters.TEXT, pay_date)],
            PAY_RECEIPT: [MessageHandler(filters.PHOTO, pay_receipt_photo),
                          MessageHandler(filters.TEXT, pay_receipt_skip),
                          CallbackQueryHandler(cb_skip_receipt, pattern="^skip_receipt$")],
        },
        fallbacks=[CommandHandler("cancel", cb_cancel)], per_message=False,
    )

    # ConvHandler: підвищення оренди
    increase_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cb_rent_increase_start, pattern=r"^increase_\w+$")],
        states={
            RENT_INCREASE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, rent_increase_amount)],
            RENT_INCREASE_NOTE:   [MessageHandler(filters.TEXT, rent_increase_note)],
        },
        fallbacks=[CommandHandler("cancel", cb_cancel)], per_message=False,
    )

    # ConvHandler: документи
    doc_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(cb_add_doc_start, pattern=r"^add_doc_\w+$")],
        states={
            DOC_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, doc_name)],
            DOC_FILE: [MessageHandler(filters.Document.ALL | filters.PHOTO, doc_file),
                       MessageHandler(filters.TEXT, doc_file_skip)],
        },
        fallbacks=[CommandHandler("cancel", cb_cancel)], per_message=False,
    )

    # Реєстрація
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("remind", cmd_remind))
    app.add_handler(CommandHandler("forecast", cmd_forecast))
    app.add_handler(add_conv)
    app.add_handler(pay_conv)
    app.add_handler(increase_conv)
    app.add_handler(doc_conv)

    app.add_handler(CallbackQueryHandler(cb_choose_pay,       pattern=r"^choose_pay_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_quickpay,         pattern=r"^quickpay_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_quickpay_confirm, pattern=r"^quickpay_confirm_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_points_detail,    pattern=r"^points_detail_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_unpay,            pattern=r"^unpay_\w+_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_show_gps,         pattern=r"^show_gps_\w+$"))
    app.add_handler(CallbackQueryHandler(cb_export,           pattern=r"^export_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_sel_month,        pattern=r"^selmonth_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_sel_year,         pattern=r"^selyear_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_sel_month_nav,    pattern="^sel_month_nav$"))
    app.add_handler(CallbackQueryHandler(cb_compare,          pattern=r"^compare_\d+_\d+$"))
    app.add_handler(CallbackQueryHandler(cb_rent_increase_list, pattern="^rent_increase_list$"))
    app.add_handler(CallbackQueryHandler(cb_docs_list,        pattern="^docs_list$"))
    app.add_handler(CallbackQueryHandler(cb_delete_point_list,pattern="^delete_point_list$"))
    app.add_handler(CallbackQueryHandler(cb_del_confirm,      pattern=r"^del_confirm_\w+$"))
    app.add_handler(CallbackQueryHandler(cb_del_do,           pattern=r"^del_do_\w+$"))
    app.add_handler(CallbackQueryHandler(cb_partners_menu,    pattern="^partners_menu$"))
    app.add_handler(CallbackQueryHandler(cb_settings_back,    pattern="^settings_back$"))
    app.add_handler(CallbackQueryHandler(cb_cancel,           pattern="^cancel$"))
    app.add_handler(CallbackQueryHandler(cb_noop,             pattern="^noop$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    # Scheduler jobs
    job_queue = app.job_queue
    job_queue.run_daily(job_daily_reminder, time=dtime(9, 0))   # кожен день о 09:00
    job_queue.run_daily(job_monthly_report, time=dtime(8, 0))   # 1-го числа о 08:00

    print("🤖 Rental Bot Enhanced — started!")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
